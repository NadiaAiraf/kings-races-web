#!/usr/bin/env python3
"""
Extract golden data from the master xlsx source of truth for all 29 team counts
(4 through 32) and produce a combined JSON fixture file.

Output: src/domain/cheatSheets/__fixtures__/goldenData.json

Per D-01: Uses Python with openpyxl.
Per D-02: Single combined JSON file keyed by team count string.
Per D-03: Seed maps extracted from column C formulas ('Enter Teams'!B{row}).
Per D-04: 32-team seedMap hardcoded (no formulas in xlsx).

Known xlsx quirks handled:
- '0' (zero) used instead of 'O' in R2 letter races
- Trailing asterisks on race entries
- Double spaces in R2 race values
- 4-6 teams have no R2 round
"""

import json
import os
import re
import sys

import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

# Search for the xlsx in several locations
XLSX_CANDIDATES = [
    os.path.join(PROJECT_ROOT, 'reference', 'cheat-sheets-v1.4.xlsx'),
    # If running from a worktree, check the main repo
    os.path.join(PROJECT_ROOT, '..', '..', 'reference', 'cheat-sheets-v1.4.xlsx'),
    # Absolute path to main repo
    '/Users/aidanfaria/kings_races_web/reference/cheat-sheets-v1.4.xlsx',
]

OUTPUT_PATH = os.path.join(
    PROJECT_ROOT, 'src', 'domain', 'cheatSheets', '__fixtures__', 'goldenData.json'
)

# Per D-04: 32-team seedMap derived from 25-31 pattern (no xlsx formulas)
HARDCODED_32_SEED_MAP = [
    1, 11, 21, 31, 41, 51, 61, 71,
    74, 64, 54, 44, 34, 24, 14, 4,
    2, 12, 22, 32, 42, 52, 62, 72,
    63, 53, 43, 33, 23, 13, 3, 73,
]


def find_xlsx():
    """Find the xlsx file in known locations."""
    for path in XLSX_CANDIDATES:
        resolved = os.path.realpath(path)
        if os.path.isfile(resolved):
            return resolved
    print("ERROR: Cannot find cheat-sheets-v1.4.xlsx in any of:")
    for p in XLSX_CANDIDATES:
        print(f"  {p}")
    sys.exit(1)


def extract_seed_map(ws, team_count):
    """Extract seed-to-slot mapping from column C formulas.

    Column C cells contain formulas like ='Enter Teams'!B{row}
    where row 3 = seed 1, row 4 = seed 2, etc.
    Column B of the same row contains the slot number.

    Returns list where index i = slot for seed (i+1).
    """
    if team_count == 32:
        print(f"  WARNING: 32 Teams -- using hardcoded seedMap (no xlsx formulas)")
        return list(HARDCODED_32_SEED_MAP)

    seed_slot_pairs = []
    max_row = ws.max_row or 100

    for row in range(1, max_row + 1):
        c_val = ws[f'C{row}'].value
        b_val = ws[f'B{row}'].value

        if c_val and isinstance(c_val, str) and "'Enter Teams'" in c_val:
            m = re.search(r"'Enter Teams'[!.]B(\d+)", c_val)
            if m and isinstance(b_val, (int, float)):
                enter_row = int(m.group(1))
                seed = enter_row - 2  # Row 3 = seed 1, row 4 = seed 2, ...
                slot = int(b_val)
                if seed >= 1:
                    seed_slot_pairs.append((seed, slot))

    # Sort by seed number, return slots in seed order
    seed_slot_pairs.sort(key=lambda x: x[0])
    return [slot for _, slot in seed_slot_pairs]


def extract_r1_races_from_col(ws, col, max_row):
    """Extract R1 race matchups from a specific column."""
    races = []
    race_num = 1
    for row in range(1, max_row + 1):
        val = ws[f'{col}{row}'].value
        if val and isinstance(val, str):
            cleaned = val.strip().rstrip('* ')
            m = re.match(r'^(\d+)\s*V\s*(\d+)$', cleaned)
            if m:
                races.append({
                    'raceNum': race_num,
                    'homeSlot': int(m.group(1)),
                    'awaySlot': int(m.group(2)),
                })
                race_num += 1
    return races


def extract_r1_races(ws):
    """Extract R1 race matchups, checking columns J and Q.

    4-6 teams use column Q; 7+ teams use column J.
    """
    max_row = ws.max_row or 100
    for col in ['J', 'Q']:
        races = extract_r1_races_from_col(ws, col, max_row)
        if races:
            return races
    return []


def extract_r2_races_from_col(ws, col, max_row):
    """Extract R2 race matchups from a specific column."""
    races = []
    race_num = 1
    for row in range(1, max_row + 1):
        val = ws[f'{col}{row}'].value
        if val and isinstance(val, str):
            cleaned = val.strip().rstrip('* ')
            # Normalize multiple spaces to single space
            cleaned = re.sub(r'\s+', ' ', cleaned)
            m = re.match(r'^([A-Za-z0-9]+)\s+V\s+([A-Za-z0-9]+)$', cleaned)
            if m:
                home = m.group(1).strip()
                away = m.group(2).strip()
                # Fix known typo: '0' -> 'O'
                if home == '0':
                    home = 'O'
                if away == '0':
                    away = 'O'
                races.append({
                    'raceNum': race_num,
                    'homeLetter': home,
                    'awayLetter': away,
                })
                race_num += 1
    return races


def extract_r2_races(ws):
    """Extract R2 race matchups, checking columns T and U.

    Most sheets use column T; 11 Teams uses column U.
    Handles '0' -> 'O' substitution (known xlsx quirk).
    Normalizes whitespace.
    """
    max_row = ws.max_row or 150
    for col in ['T', 'U']:
        races = extract_r2_races_from_col(ws, col, max_row)
        if races:
            return races
    return []


def extract_sheet(wb, team_count):
    """Extract golden data for a single team count sheet."""
    sheet_name = f'{team_count} Teams'
    ws = wb[sheet_name]

    # Seed map (uses formula workbook)
    seed_map = extract_seed_map(ws, team_count)

    # R1 races
    r1_races = extract_r1_races(ws)

    # R2 races (empty for 4-6 teams per pitfall 4)
    if team_count >= 7:
        r2_races = extract_r2_races(ws)
    else:
        r2_races = []

    return {
        'teamCount': team_count,
        'seedMap': seed_map,
        'r1Races': r1_races,
        'r2Races': r2_races,
    }


def main():
    xlsx_path = find_xlsx()
    print(f'Loading workbook: {xlsx_path}')
    print(f'  (data_only=False to read formulas)')

    # Load with data_only=False to read formulas for seed map extraction
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    # Create output directory
    output_dir = os.path.dirname(OUTPUT_PATH)
    os.makedirs(output_dir, exist_ok=True)

    golden_data = {}
    errors = []

    for team_count in range(4, 33):
        sheet_name = f'{team_count} Teams'
        if sheet_name not in wb.sheetnames:
            errors.append(f'{sheet_name}: Sheet not found!')
            continue

        print(f'Processing: {sheet_name}')
        data = extract_sheet(wb, team_count)

        # Validation
        seed_len = len(data['seedMap'])
        r1_len = len(data['r1Races'])
        r2_len = len(data['r2Races'])

        print(f'  seedMap: {seed_len}, R1 races: {r1_len}, R2 races: {r2_len}')

        if seed_len != team_count:
            msg = f'{sheet_name}: seedMap length {seed_len} != teamCount {team_count}'
            errors.append(msg)
            print(f'  WARNING: {msg}')

        if seed_len == 0:
            msg = f'{sheet_name}: seedMap is EMPTY!'
            errors.append(msg)
            print(f'  WARNING: {msg}')

        if r1_len == 0:
            msg = f'{sheet_name}: No R1 races extracted!'
            errors.append(msg)
            print(f'  WARNING: {msg}')

        if team_count >= 7 and r2_len == 0:
            msg = f'{sheet_name}: No R2 races extracted (expected for 7+ teams)!'
            errors.append(msg)
            print(f'  WARNING: {msg}')

        golden_data[str(team_count)] = data

    # Write output
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(golden_data, f, indent=2)
    print(f'\nWrote golden data to: {OUTPUT_PATH}')

    # Summary
    print(f'\nSummary: {len(golden_data)} entries')
    for tc_str in sorted(golden_data.keys(), key=int):
        d = golden_data[tc_str]
        print(f'  {tc_str} teams: seedMap={len(d["seedMap"])}, R1={len(d["r1Races"])}, R2={len(d["r2Races"])}')

    if errors:
        print(f'\n=== ERRORS ({len(errors)}) ===')
        for e in errors:
            print(f'  {e}')
        sys.exit(1)
    else:
        print('\nAll 29 sheets extracted successfully!')


if __name__ == '__main__':
    main()
