#!/usr/bin/env python3
"""
Extract all 29 cheat sheet race order sequences from the existing xlsx spreadsheet
into typed TypeScript const files.

Per D-01: Extracts programmatically from the xlsx file.
Per D-02: Only one set of 29 sequences (identical across disciplines).
Per D-03: Preserves full tournament structure (groups, R1, R2, finals).
Per D-04: Group assignments extracted exactly (intentionally seeded).
Per D-05: Finals structures vary by team count, extracted per sheet.

Known spreadsheet quirks handled:
- Lowercase 'v' used as team letter when 'V' would conflict with 'V' separator (22+ teams)
- '0' (zero) used instead of 'O' in "N V 0" (15 Teams T14)
- Trailing asterisk on race entries (e.g., "51 V 52*" in 17 Teams)
- COUNTA-based race counts include header rows (subtract 1 for actual race count)
- Group labels in column A missing for some large sheets; use column C "Group X" headers
- 4-6 teams have no column A group labels; single group inferred from B column
"""

import openpyxl
import re
import os
from openpyxl.utils import get_column_letter

XLSX_PATH = '/Users/aidanfaria/Downloads/Master Copy DO NOT USE v1.4 - cheat sheets - auto populating.xlsx'
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'src', 'domain', 'cheatSheets')


def extract_groups(ws, max_row):
    """Extract group letter -> team slot numbers.
    Uses column A labels when available, falls back to column C 'Group X' headers
    with team slots from column B."""
    groups = {}

    # Try column A labels first
    for row in range(1, max_row + 1):
        a_val = ws[f'A{row}'].value
        b_val = ws[f'B{row}'].value
        if a_val and isinstance(a_val, str) and len(a_val) == 1 and a_val.isalpha() and a_val.isupper():
            if b_val and isinstance(b_val, (int, float)):
                slot = int(b_val)
                groups.setdefault(a_val, []).append(slot)

    # Check if column C has more groups than column A (partial A labels)
    c_group_count = 0
    for row in range(1, max_row + 1):
        c_val = ws[f'C{row}'].value
        if c_val and isinstance(c_val, str) and c_val.startswith('Group '):
            letter = c_val.replace('Group ', '').strip()
            if len(letter) == 1 and letter.isalpha():
                c_group_count += 1

    if groups and c_group_count <= len(groups):
        return groups

    # Use column C "Group X" headers (either as fallback or because A is incomplete)
    group_headers = []  # (row, letter)
    for row in range(1, max_row + 1):
        c_val = ws[f'C{row}'].value
        if c_val and isinstance(c_val, str) and c_val.startswith('Group '):
            letter = c_val.replace('Group ', '').strip()
            if len(letter) == 1 and letter.isalpha():
                group_headers.append((row, letter))

    if not group_headers:
        return groups

    # For each group header, collect team slots from B column until next group or gap
    for i, (start_row, letter) in enumerate(group_headers):
        end_row = group_headers[i + 1][0] if i + 1 < len(group_headers) else max_row + 1
        slots = []
        for row in range(start_row, end_row):
            b_val = ws[f'B{row}'].value
            if isinstance(b_val, (int, float)) and b_val > 0:
                slots.append(int(b_val))
        if slots:
            groups[letter] = slots

    return groups


def extract_races_from_column(ws, col_letter, max_row):
    """Extract 'N V M' numeric race pairs from a column.
    Handles trailing asterisks ('51 V 52*') and missing spaces ('4 V1')."""
    races = []
    for row in range(1, max_row + 1):
        val = ws[f'{col_letter}{row}'].value
        if val and isinstance(val, str):
            # Strip trailing non-digit characters (asterisks, spaces)
            cleaned = val.strip().rstrip('* ')
            match = re.match(r'^(\d+)\s*V\s*(\d+)$', cleaned)
            if match:
                races.append((int(match.group(1)), int(match.group(2))))
    return races


def extract_letter_races_from_column(ws, col_letter, max_row):
    """Extract letter-based race pairs like 'A V B', 'CC V DD', 'U V v' from a column.
    Handles lowercase 'v' (used when letter V would conflict with separator)
    and '0' (zero) used as typo for 'O'."""
    races = []
    for row in range(1, max_row + 1):
        val = ws[f'{col_letter}{row}'].value
        if val and isinstance(val, str):
            # Strip trailing asterisks and whitespace
            cleaned = val.strip().rstrip('* ')
            # Match letter races including lowercase and digits (for 0/O typo)
            match = re.match(r'^([A-Za-z0-9]+)\s+V\s+([A-Za-z0-9]+)\s*$', cleaned)
            if match:
                home = match.group(1)
                away = match.group(2)
                # Fix known typos: '0' -> 'O'
                if home == '0':
                    home = 'O'
                if away == '0':
                    away = 'O'
                races.append((home, away))
    return races


def find_r1_race_column(ws, max_row):
    """Find which column contains Round 1 numeric race order."""
    for col in ['J', 'Q']:
        races = extract_races_from_column(ws, col, max_row)
        if races:
            return col, races
    return None, []


def find_r2_race_column(ws, max_row):
    """Find which column contains Round 2 letter-based race order."""
    for col in ['T', 'U']:
        races = extract_letter_races_from_column(ws, col, max_row)
        if races:
            return col, races
    return None, []


def extract_r2_seeding(ws, max_row, team_count):
    """Extract Round 2 seeding: position code, letter, label.
    Column layout varies: K/L/M or L/M/N depending on sheet."""
    seedings = []

    # Try K/L/M first (most common for sheets with position codes)
    for row in range(1, max_row + 1):
        k_val = ws[f'K{row}'].value
        l_val = ws[f'L{row}'].value
        m_val = ws[f'M{row}'].value

        if k_val and l_val and m_val:
            k_str = str(k_val).strip()
            l_str = str(l_val).strip()
            m_str = str(m_val).strip()
            # Position code like 'A1', 'B2', etc.
            if re.match(r'^[A-Z]\d$', k_str) and re.match(r'^[A-Za-z]+$', l_str) and len(m_str) > 2:
                seedings.append({
                    'positionCode': k_str,
                    'letter': l_str,
                    'label': m_str,
                })

    if seedings:
        return seedings

    # Try L/M/N (used by 11 Teams)
    for row in range(1, max_row + 1):
        l_val = ws[f'L{row}'].value
        m_val = ws[f'M{row}'].value
        n_val = ws[f'N{row}'].value

        if l_val and m_val and n_val:
            l_str = str(l_val).strip()
            m_str = str(m_val).strip()
            n_str = str(n_val).strip()
            if re.match(r'^[A-Z]\d$', l_str) and re.match(r'^[A-Za-z]+$', m_str) and len(n_str) > 2:
                seedings.append({
                    'positionCode': l_str,
                    'letter': m_str,
                    'label': n_str,
                })

    if seedings:
        return seedings

    # For sheets without position codes (like 17+ Teams), try L/M only
    for row in range(1, max_row + 1):
        l_val = ws[f'L{row}'].value
        m_val = ws[f'M{row}'].value

        if l_val and m_val:
            l_str = str(l_val).strip()
            m_str = str(m_val).strip()
            # Match single or multi-letter references (A, B, ..., AA, BB, v)
            if re.match(r'^[A-Za-z]+$', l_str) and len(m_str) > 2:
                # Skip group headers and non-seeding values
                if m_str.startswith('Group ') or m_str in ['Race Order', 'ROUND TWO']:
                    continue
                seedings.append({
                    'positionCode': '',
                    'letter': l_str,
                    'label': m_str,
                })

    return seedings


def extract_r2_groups(ws, max_row, seedings, r2_races, team_count):
    """Extract Round 2 group structure from the sheet.
    Groups are labeled with Roman numerals (I, II, III, etc.)."""
    groups = []

    # Determine which column has the R2 group labels
    group_col = None
    for col in ['M', 'N']:
        for row in range(1, max_row + 1):
            val = ws[f'{col}{row}'].value
            if val and isinstance(val, str) and val.startswith('Group '):
                # Check if it's a Roman numeral group (R2) not a letter group (R1)
                suffix = val.replace('Group ', '').strip()
                if suffix in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII']:
                    group_col = col
                    break
        if group_col:
            break

    if not group_col:
        return []

    # Scan for R2 group headers
    group_labels = []
    group_rows = []
    for row in range(1, max_row + 1):
        val = ws[f'{group_col}{row}'].value
        if val and isinstance(val, str) and val.startswith('Group '):
            suffix = val.replace('Group ', '').strip()
            if suffix in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII']:
                group_labels.append(suffix)
                group_rows.append(row)

    if not group_labels:
        return []

    # Build seeding map: letter -> seeding entry
    seeding_map = {}
    for s in seedings:
        seeding_map[s['letter']] = s

    # Find which column has the seeding letters
    seeding_letter_col = None
    for col in ['L', 'M']:
        for row in range(1, max_row + 1):
            val = ws[f'{col}{row}'].value
            if val and isinstance(val, str) and val.strip() in seeding_map:
                seeding_letter_col = col
                break
        if seeding_letter_col:
            break

    if not seeding_letter_col:
        return []

    # Build row -> letter mapping
    seeding_rows = {}
    for row in range(1, max_row + 1):
        val = ws[f'{seeding_letter_col}{row}'].value
        if val and isinstance(val, str) and val.strip() in seeding_map:
            seeding_rows[row] = val.strip()

    # Assign seedings to groups based on row positions
    for i, (label, start_row) in enumerate(zip(group_labels, group_rows)):
        end_row = group_rows[i + 1] if i + 1 < len(group_rows) else max_row + 1
        group_seedings = []
        for row in range(start_row, end_row):
            if row in seeding_rows:
                letter = seeding_rows[row]
                if letter in seeding_map:
                    group_seedings.append(seeding_map[letter])

        # Find R2 races where both teams are in this group
        group_letters = set(s['letter'] for s in group_seedings)
        group_races = []
        race_num = 1
        for home, away in r2_races:
            if home in group_letters and away in group_letters:
                group_races.append({
                    'raceNum': race_num,
                    'homeLetter': home,
                    'awayLetter': away,
                })
                race_num += 1

        groups.append({
            'groupNum': label,
            'seedingEntries': group_seedings,
            'races': group_races,
        })

    return groups


def extract_finals(ws, max_row, team_count):
    """Extract finals matchups from the sheet.
    Format and location varies significantly per team count."""
    finals = []

    if team_count <= 6:
        # For 4-6 teams, look for placement labels like "1st/2nd", "3rd/4th"
        for row in range(1, max_row + 1):
            for label_col in ['Y', 'AA']:
                label_val = ws[f'{label_col}{row}'].value
                if label_val and isinstance(label_val, str) and '/' in label_val and any(c.isdigit() for c in label_val):
                    home_ref = None
                    away_ref = None
                    for h_col in ['T']:
                        h_val = ws[f'{h_col}{row}'].value
                        if h_val and isinstance(h_val, str) and len(h_val.strip()) > 3:
                            home_ref = h_val.strip()
                    for a_col in ['X']:
                        a_val = ws[f'{a_col}{row}'].value
                        if a_val and isinstance(a_val, str) and len(a_val.strip()) > 3:
                            away_ref = a_val.strip()
                    if home_ref and away_ref:
                        finals.append({
                            'label': label_val.strip(),
                            'homeRef': home_ref,
                            'awayRef': away_ref,
                        })
        return finals

    # For 7+ teams, try labeled finals first (AA or AB column has labels)
    for row in range(1, max_row + 1):
        for label_col in ['AA', 'AB', 'Y']:
            label_val = ws[f'{label_col}{row}'].value
            if label_val and isinstance(label_val, str) and '/' in label_val and any(c.isdigit() for c in label_val):
                home_ref = None
                away_ref = None

                for h_col in ['V', 'W']:
                    h_val = ws[f'{h_col}{row}'].value
                    if h_val and isinstance(h_val, str) and ('Group' in h_val or 'Winner' in h_val or 'Loser' in h_val or 'Second' in h_val or 'Third' in h_val or 'Runner' in h_val):
                        home_ref = h_val.strip()
                        break

                for a_col in ['Z', 'AA', 'X']:
                    if a_col == label_col:
                        continue
                    a_val = ws[f'{a_col}{row}'].value
                    if a_val and isinstance(a_val, str) and ('Group' in a_val or 'Winner' in a_val or 'Loser' in a_val or 'Second' in a_val or 'Third' in a_val or 'Runner' in a_val):
                        away_ref = a_val.strip()
                        break

                if home_ref and away_ref:
                    finals.append({
                        'label': label_val.strip(),
                        'homeRef': home_ref,
                        'awayRef': away_ref,
                    })

    if finals:
        return finals

    # Fallback: look for "V" separator in column X or Y without explicit labels
    for row in range(1, max_row + 1):
        for v_col in ['X', 'Y']:
            v_val = ws[f'{v_col}{row}'].value
            if v_val and isinstance(v_val, str) and v_val.strip() in ['V', '*']:
                home_ref = None
                away_ref = None

                for h_col in ['V', 'W']:
                    h_val = ws[f'{h_col}{row}'].value
                    if h_val and isinstance(h_val, str) and len(h_val.strip()) > 3 and h_val.strip() != 'FINAL ROUND':
                        home_ref = h_val.strip()
                        break

                for a_col in ['Z', 'AA']:
                    a_val = ws[f'{a_col}{row}'].value
                    if a_val and isinstance(a_val, str) and len(a_val.strip()) > 3:
                        away_ref = a_val.strip()
                        break

                if home_ref and away_ref:
                    finals.append({
                        'label': '',
                        'homeRef': home_ref,
                        'awayRef': away_ref,
                    })

    # Assign placement labels (last = 1st/2nd, etc.)
    if finals and not any(f['label'] for f in finals):
        n = len(finals)
        for i, f in enumerate(finals):
            pos = (n - i) * 2
            f['label'] = _ordinal(pos - 1) + '/' + _ordinal(pos)

    return finals


def _ordinal(n):
    """Convert number to ordinal string: 1st, 2nd, 3rd, 4th, etc."""
    if n % 10 == 1 and n % 100 != 11:
        return f'{n}st'
    elif n % 10 == 2 and n % 100 != 12:
        return f'{n}nd'
    elif n % 10 == 3 and n % 100 != 13:
        return f'{n}rd'
    else:
        return f'{n}th'


def find_race_count(ws, col_letter, max_row):
    """Find the race count value from a 'races:' label.
    Note: spreadsheet uses COUNTA which includes headers, so actual race count = value - 1."""
    for row in range(1, max_row + 1):
        val = ws[f'{col_letter}{row}'].value
        if val and isinstance(val, str) and 'race' in val.lower():
            # Check next row for the count
            next_val = ws[f'{col_letter}{row + 1}'].value
            if isinstance(next_val, (int, float)):
                return int(next_val)
            # Check same row in adjacent columns
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            for offset in range(1, 4):
                adj_col = get_column_letter(col_idx + offset)
                adj_val = ws[f'{adj_col}{row}'].value
                if isinstance(adj_val, (int, float)):
                    return int(adj_val)
    return None


def escape_ts_string(s):
    return s.replace("'", "\\'")


def generate_ts_file(team_count, data):
    """Generate TypeScript file content for a cheat sheet."""
    lines = []
    lines.append("import type { TournamentStructure } from '../types';")
    lines.append('')
    lines.append(f'export const TEAMS_{team_count}: TournamentStructure = {{')
    lines.append(f'  teamCount: {team_count},')

    # Groups
    lines.append('  groups: [')
    for g in data['groups']:
        slots_str = ', '.join(str(s) for s in g['teamSlots'])
        lines.append(f"    {{ letter: '{g['letter']}', teamSlots: [{slots_str}] }},")
    lines.append('  ],')

    # Round One Races
    lines.append('  roundOneRaces: [')
    for i, (home, away) in enumerate(data['roundOneRaces']):
        lines.append(f'    {{ raceNum: {i + 1}, homeSlot: {home}, awaySlot: {away} }},')
    lines.append('  ],')

    # Round Two Groups (if applicable)
    if data.get('roundTwoGroups'):
        lines.append('  roundTwoGroups: [')
        for group in data['roundTwoGroups']:
            lines.append('    {')
            lines.append(f"      groupNum: '{group['groupNum']}',")
            lines.append('      seedingEntries: [')
            for s in group['seedingEntries']:
                pc = escape_ts_string(s['positionCode'])
                lt = escape_ts_string(s['letter'])
                lb = escape_ts_string(s['label'])
                lines.append(f"        {{ positionCode: '{pc}', letter: '{lt}', label: '{lb}' }},")
            lines.append('      ],')
            lines.append('      races: [')
            for r in group['races']:
                hl = escape_ts_string(r['homeLetter'])
                al = escape_ts_string(r['awayLetter'])
                lines.append(f"        {{ raceNum: {r['raceNum']}, homeLetter: '{hl}', awayLetter: '{al}' }},")
            lines.append('      ],')
            lines.append('    },')
        lines.append('  ],')

    # Finals
    lines.append('  finals: [')
    for f in data['finals']:
        lb = escape_ts_string(f['label'])
        hr = escape_ts_string(f['homeRef'])
        ar = escape_ts_string(f['awayRef'])
        lines.append(f"    {{ label: '{lb}', homeRef: '{hr}', awayRef: '{ar}' }},")
    lines.append('  ],')

    # Race counts
    lines.append(f"  roundOneRaceCount: {data['roundOneRaceCount']},")
    if data.get('roundTwoRaceCount') is not None:
        lines.append(f"  roundTwoRaceCount: {data['roundTwoRaceCount']},")

    lines.append('};')
    lines.append('')

    return '\n'.join(lines)


def extract_sheet(wb, team_count):
    """Extract all data from a single cheat sheet."""
    name = f'{team_count} Teams'
    ws = wb[name]
    max_row = ws.max_row or 100

    # Extract groups
    groups_dict = extract_groups(ws, max_row)

    # For 4-6 teams with no group labels, infer single group from B column
    if not groups_dict and team_count <= 6:
        slots = []
        for row in range(1, max_row + 1):
            b_val = ws[f'B{row}'].value
            if isinstance(b_val, (int, float)) and b_val > 0:
                slots.append(int(b_val))
        if slots:
            groups_dict = {'A': sorted(slots)}

    # Convert to sorted list
    groups = []
    for letter in sorted(groups_dict.keys()):
        groups.append({
            'letter': letter,
            'teamSlots': groups_dict[letter],
        })

    # Extract Round 1 races
    r1_col, r1_races = find_r1_race_column(ws, max_row)
    r1_count = len(r1_races)

    # Note: Spreadsheet COUNTA includes headers (varying count).
    # Race counts are verified by the actual data extraction, not the COUNTA formula.

    data = {
        'groups': groups,
        'roundOneRaces': r1_races,
        'roundOneRaceCount': r1_count,
    }

    # For 7+ teams, extract Round 2
    if team_count >= 7:
        r2_col, r2_races = find_r2_race_column(ws, max_row)
        r2_count = len(r2_races)

        # Verify R2 count (COUNTA includes "Race Order" header + group headers)
        if r2_col:
            spreadsheet_r2 = find_race_count(ws, r2_col, max_row)
            if spreadsheet_r2 is None:
                spreadsheet_r2 = find_race_count(ws, 'U', max_row)
            # Don't warn here -- COUNTA includes varying number of headers

        # Extract R2 seeding
        seedings = extract_r2_seeding(ws, max_row, team_count)

        # Extract R2 groups
        r2_groups = extract_r2_groups(ws, max_row, seedings, r2_races, team_count)

        data['roundTwoGroups'] = r2_groups
        data['roundTwoRaceCount'] = r2_count

    # Extract finals
    finals = extract_finals(ws, max_row, team_count)
    data['finals'] = finals

    return data


def main():
    print(f'Loading workbook: {XLSX_PATH}')
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    errors = []
    for team_count in range(4, 33):
        name = f'{team_count} Teams'
        if name not in wb.sheetnames:
            print(f'  SKIP: {name} not found')
            continue

        print(f'Processing: {name}')
        data = extract_sheet(wb, team_count)

        # Validation
        if not data['groups']:
            errors.append(f'{name}: No groups extracted!')
        if not data['roundOneRaces']:
            errors.append(f'{name}: No R1 races extracted!')
        if team_count >= 7 and not data.get('roundTwoGroups'):
            errors.append(f'{name}: No R2 groups extracted!')
        if not data['finals']:
            print(f'  NOTE: {name} has no finals (may be optional for small team counts)')

        ts_content = generate_ts_file(team_count, data)
        output_path = os.path.join(OUTPUT_DIR, f'teams{team_count}.ts')
        with open(output_path, 'w') as f:
            f.write(ts_content)

        # Summary
        r1 = len(data['roundOneRaces'])
        r2 = data.get('roundTwoRaceCount', 0)
        g = len(data['groups'])
        r2g = len(data.get('roundTwoGroups', []))
        fin = len(data['finals'])
        print(f'  Groups: {g}, R1 races: {r1}, R2 groups: {r2g}, R2 races: {r2}, Finals: {fin}')

    if errors:
        print('\n=== ERRORS ===')
        for e in errors:
            print(f'  {e}')
    else:
        print('\nAll sheets extracted successfully!')

    print(f'\nFiles written to {OUTPUT_DIR}')


if __name__ == '__main__':
    main()
